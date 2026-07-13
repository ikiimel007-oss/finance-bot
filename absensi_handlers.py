from datetime import datetime
import os
import tempfile

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    CommandHandler,
    CallbackQueryHandler,
    ConversationHandler,
    MessageHandler,
    filters,
    ContextTypes,
)

import database as db

(
    ABSENSI_ADD_MEMBER,
    ABSENSI_CUSTOM_DATE,
    ABSENSI_CEK_TANGGAL,
    ABSENSI_PERSENTASE,
    ABSENSI_EXCEL,
) = range(20, 25)

MONTH_NAMES = [
    "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
]


def menu_button_abs():
    return [InlineKeyboardButton("🏠 Menu Utama", callback_data="menu_main")]

def wrap_keyboard_abs(buttons):
    return InlineKeyboardMarkup(buttons + [menu_button_abs()])


def absensi_menu_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📝 Check-in", callback_data="abs_checkin")],
        [InlineKeyboardButton("👥 Daftar Anggota", callback_data="abs_list_members")],
        [InlineKeyboardButton("➕ Tambah Anggota", callback_data="abs_add_member")],
        [InlineKeyboardButton("🗑 Hapus Anggota", callback_data="abs_delete_member")],
        [InlineKeyboardButton("📋 Cek Hari Ini", callback_data="abs_check_today")],
        [InlineKeyboardButton("📅 Cek Tanggal", callback_data="abs_check_date")],
        [InlineKeyboardButton("📊 Persentase", callback_data="abs_percentage")],
        [InlineKeyboardButton("🗑 Hapus Absen", callback_data="abs_delete_record")],
        [InlineKeyboardButton("📥 Download Excel", callback_data="abs_excel_month")],
    ])


def get_today():
    return datetime.now().strftime("%Y-%m-%d")

def get_now():
    return datetime.now().strftime("%H:%M:%S")


async def absensi_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    text = (
        "📋 *Menu Absensi*\n"
        "━━━━━━━━━━━━━━━━━━━━━━\n"
        "Pilih menu di bawah:"
    )
    await query.edit_message_text(text, parse_mode="Markdown", reply_markup=absensi_menu_keyboard())


# ─── CHECK-IN FLOW ──────────────────────────────────────────

async def abs_checkin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    members = db.get_members()
    if not members:
        await query.edit_message_text(
            "📭 Belum ada anggota. Tambah anggota dulu.",
            reply_markup=wrap_keyboard_abs([]),
        )
        return

    # Init session
    context.user_data["abs_selected"] = set()
    context.user_data["abs_date"] = get_today()
    context.user_data["abs_mode"] = "hadir"

    await _show_checkin_menu(update, context)


async def _show_checkin_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    sel = context.user_data.get("abs_selected", set())
    mode = context.user_data.get("abs_mode", "hadir")
    date = context.user_data.get("abs_date", get_today())
    members = db.get_members()

    buttons = []
    for m in members:
        checked = "☑" if m["id"] in sel else "☐"
        buttons.append([InlineKeyboardButton(f"{checked} {m['name']}", callback_data=f"abspick_{m['id']}")])

    mode_label = "✅ Hadir" if mode == "hadir" else "📋 Izin"
    buttons.append([InlineKeyboardButton(f"🔄 Mode: {mode_label}", callback_data="abs_toggle_mode")])
    buttons.append([InlineKeyboardButton("📅 Custom Tanggal", callback_data="abs_custom_date")])
    buttons.append([InlineKeyboardButton("✅ SAVE", callback_data="abs_save_checkin")])
    buttons.append(menu_button_abs())

    text = (
        f"📝 *Check-in* — {date}\n"
        f"Mode: {mode_label}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━\n"
        f"Pilih anggota (bisa banyak):\n"
        f"Terpilih: {len(sel)}"
    )
    await query.edit_message_text(text, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(buttons))


async def abs_pick_member(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    member_id = int(query.data.split("_")[1])
    sel = context.user_data.setdefault("abs_selected", set())
    if member_id in sel:
        sel.remove(member_id)
    else:
        sel.add(member_id)
    await _show_checkin_menu(update, context)


async def abs_toggle_mode(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    mode = context.user_data.get("abs_mode", "hadir")
    context.user_data["abs_mode"] = "izin" if mode == "hadir" else "hadir"
    await _show_checkin_menu(update, context)


async def abs_custom_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(
        "📅 Kirim tanggal check-in (format: *dd-mm-yyyy*)\n"
        "Contoh: `15-07-2026`",
        parse_mode="Markdown",
        reply_markup=wrap_keyboard_abs([]),
    )
    return ABSENSI_CUSTOM_DATE


async def abs_custom_date_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    import re
    m = re.match(r"^(\d{2})-(\d{2})-(\d{4})$", text)
    if not m:
        await update.message.reply_text("⚠️ Format salah. Gunakan: dd-mm-yyyy\nContoh: `15-07-2026`", parse_mode="Markdown")
        return ABSENSI_CUSTOM_DATE
    date = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    context.user_data["abs_date"] = date
    context.user_data["abs_selected"] = set()
    context.user_data["abs_mode"] = "hadir"

    members = db.get_members()
    if not members:
        await update.message.reply_text("📭 Belum ada anggota.")
        return ConversationHandler.END

    buttons = []
    for mm in members:
        buttons.append([InlineKeyboardButton(f"☐ {mm['name']}", callback_data=f"abspick_{mm['id']}")])
    buttons.append([InlineKeyboardButton("📅 Custom Tanggal", callback_data="abs_custom_date")])
    buttons.append([InlineKeyboardButton("✅ SAVE", callback_data="abs_save_checkin")])
    buttons.append(menu_button_abs())

    await update.message.reply_text(
        f"📝 *Check-in* — {date}\n\nPilih anggota (bisa banyak):",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(buttons),
    )
    return ConversationHandler.END


async def abs_save_checkin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    sel = context.user_data.get("abs_selected", set())
    if not sel:
        await query.answer("⚠️ Belum ada yang dipilih!")
        return

    date = context.user_data.get("abs_date", get_today())
    mode = context.user_data.get("abs_mode", "hadir")
    now = get_now()
    done = []
    skipped = []

    for member_id in sel:
        existing = db.get_absensi_record(member_id, date)
        if existing:
            member = db.get_member(member_id)
            skipped.append(member["name"] if member else "?")
        else:
            db.add_absensi(member_id, date, now, mode)
            member = db.get_member(member_id)
            done.append(member["name"] if member else "?")

    mode_label = "📋 Izin" if mode == "izin" else "✅ Hadir"
    msg = f"{mode_label} {date}:\n" + "\n".join(done)
    if skipped:
        msg += f"\n\n⚠️ Sudah terdaftar: {', '.join(skipped)}"

    # Clear session
    context.user_data.pop("abs_selected", None)
    context.user_data.pop("abs_date", None)
    context.user_data.pop("abs_mode", None)

    await query.answer(f"✅ {len(done)} tersimpan!")
    await query.edit_message_text(msg, reply_markup=absensi_menu_keyboard())


# ─── ANGGOTA ──────────────────────────────────────────────

async def abs_list_members(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    members = db.get_members()
    if not members:
        text = "📭 Belum ada anggota terdaftar."
    else:
        text = f"👥 *Daftar Anggota ({len(members)})*\n" + "━━━━━━━━━━━━━━━━━━━━━━\n"
        text += "\n".join(f"{i+1}. {m['name']}" for i, m in enumerate(members))
    await query.edit_message_text(text, parse_mode="Markdown", reply_markup=wrap_keyboard_abs([]))


async def abs_add_member_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(
        "✏️ Kirim nama anggota baru\n\n"
        "Bisa banyak nama, pisahkan dengan *enter* (baris baru).",
        parse_mode="Markdown",
        reply_markup=wrap_keyboard_abs([]),
    )
    return ABSENSI_ADD_MEMBER


async def abs_add_member_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    names = [n.strip() for n in text.split("\n") if n.strip()]
    if not names:
        await update.message.reply_text("⚠️ Nama tidak boleh kosong.")
        return ABSENSI_ADD_MEMBER

    added, skipped = db.add_members_bulk(names)
    msg = ""
    if added:
        msg += f"✅ {len(added)} ditambahkan: {', '.join(added)}"
    if skipped:
        msg += (msg and "\n") + f"⚠️ {len(skipped)} sudah ada: {', '.join(skipped)}"
    if not msg:
        msg = "⚠️ Tidak ada nama valid."

    await update.message.reply_text(msg, reply_markup=absensi_menu_keyboard())
    return ConversationHandler.END


async def abs_delete_member_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    members = db.get_members()
    if not members:
        await query.edit_message_text("📭 Belum ada anggota.", reply_markup=wrap_keyboard_abs([]))
        return

    buttons = [[InlineKeyboardButton(f"🗑 {m['name']}", callback_data=f"absdel_{m['id']}")] for m in members]
    buttons.append(menu_button_abs())
    await query.edit_message_text(
        "🗑 *Pilih anggota yang ingin dihapus:*",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def abs_delete_member_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    member_id = int(query.data.split("_")[1])
    member = db.get_member(member_id)
    if not member:
        await query.edit_message_text("❌ Anggota tidak ditemukan.", reply_markup=wrap_keyboard_abs([]))
        return
    db.delete_member(member_id)
    await query.edit_message_text(
        f"✅ {member['name']} berhasil dihapus.",
        reply_markup=absensi_menu_keyboard(),
    )


# ─── CEK ABSENSI ──────────────────────────────────────────

async def abs_check_today(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    today = get_today()
    data = db.get_absensi_by_date(today)
    if not data:
        text = f"📋 *Absensi Hari Ini ({today})*\n\n📭 Belum ada anggota."
    else:
        lines = []
        for d in data:
            if d.get("status") == "izin":
                lines.append(f"- {d['name']}: 📋 Izin")
            elif d.get("check_in"):
                lines.append(f"- {d['name']}: ✅ {d['check_in']}")
            else:
                lines.append(f"- {d['name']}: ❌ Tidak hadir")
        text = f"📋 *Absensi Hari Ini ({today})*\n\n" + "\n".join(lines)
    await query.edit_message_text(text, parse_mode="Markdown", reply_markup=wrap_keyboard_abs([]))


async def abs_check_date_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(
        "📅 Kirim tanggal (format: *dd-mm-yyyy*)\nContoh: `12-07-2026`",
        parse_mode="Markdown",
        reply_markup=wrap_keyboard_abs([]),
    )
    return ABSENSI_CEK_TANGGAL


async def abs_check_date_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    import re
    m = re.match(r"^(\d{2})-(\d{2})-(\d{4})$", text)
    if not m:
        await update.message.reply_text("⚠️ Format salah. Gunakan: dd-mm-yyyy")
        return ABSENSI_CEK_TANGGAL
    date = f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    data = db.get_absensi_by_date(date)
    if not data:
        msg = f"📋 *Absensi {text}*\n\n📭 Tidak ada data."
    else:
        lines = []
        for d in data:
            if d.get("status") == "izin":
                lines.append(f"- {d['name']}: 📋 Izin")
            elif d.get("check_in"):
                lines.append(f"- {d['name']}: ✅ {d['check_in']}")
            else:
                lines.append(f"- {d['name']}: ❌ Tidak hadir")
        msg = f"📋 *Absensi {text}*\n\n" + "\n".join(lines)
    await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=absensi_menu_keyboard())
    return ConversationHandler.END


# ─── PERSENTASE ──────────────────────────────────────────

async def abs_percentage_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(
        "📊 Kirim bulan dan tahun\nContoh: `7 2026`",
        parse_mode="Markdown",
        reply_markup=wrap_keyboard_abs([]),
    )
    return ABSENSI_PERSENTASE


async def abs_percentage_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    parts = text.split()
    try:
        bulan = int(parts[0])
        tahun = int(parts[1]) if len(parts) > 1 else datetime.now().year
        if bulan < 1 or bulan > 12:
            raise ValueError
    except (ValueError, IndexError):
        await update.message.reply_text("⚠️ Format: bulan tahun, contoh: `7 2026`", parse_mode="Markdown")
        return ABSENSI_PERSENTASE

    hari_aktif = db.get_attendance_dates(tahun, bulan)
    hari_kerja = len(hari_aktif)
    summary = db.get_month_attendance_summary(tahun, bulan)

    if not summary:
        await update.message.reply_text(
            f"📊 *Persentase Kehadiran {MONTH_NAMES[bulan]} {tahun}*\n\n📭 Belum ada data.",
            parse_mode="Markdown",
            reply_markup=absensi_menu_keyboard(),
        )
        return ConversationHandler.END

    lines = []
    for s in summary:
        hadir = s["hadir"] if not hasattr(s, "get") else s.get("hadir", 0)
        izin = s["izin"] if not hasattr(s, "get") else s.get("izin", 0)
        nama = s["name"] if not hasattr(s, "get") else s.get("name", "?")
        pct = (hadir / hari_kerja * 100) if hari_kerja else 0
        # Riwayat
        mid = s["id"] if not hasattr(s, "get") else s.get("id", 0)
        riwayat = db.get_member_month_absensi(mid, tahun, bulan)
        riwayat_str = ""
        if riwayat:
            parts_r = []
            for r in riwayat:
                day = r["date"][-2:] if not hasattr(r, "get") else r.get("date", "")[-2:]
                if r.get("status") == "izin":
                    parts_r.append(f"{day}(izin)")
                else:
                    parts_r.append(f"{day} ({r.get('check_in', '?')})")
            if parts_r:
                riwayat_str = "\n   📆 " + ", ".join(parts_r)
        lines.append(f"- {nama}: {hadir}/{hari_kerja} ({pct:.1f}%){riwayat_str}")

    msg = f"📊 *Persentase Kehadiran {MONTH_NAMES[bulan]} {tahun}*\n(Hari aktif: {hari_kerja})\n\n" + "\n".join(lines)
    await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=absensi_menu_keyboard())
    return ConversationHandler.END


# ─── HAPUS ABSEN ──────────────────────────────────────────

async def abs_delete_record_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = db.get_recent_absensi(30)
    if not data:
        await query.edit_message_text("📭 Tidak ada data absen.", reply_markup=wrap_keyboard_abs([]))
        return

    buttons = []
    for d in data:
        label = f"🗑 #{d['id']} {d['name']} ({d['date']})"
        buttons.append([InlineKeyboardButton(label, callback_data=f"absdelrec_{d['id']}")])
    buttons.append(menu_button_abs())
    await query.edit_message_text(
        "🗑 *Pilih data absen yang ingin dihapus:*",
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def abs_delete_record_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    record_id = int(query.data.split("_")[1])
    if db.delete_absensi_record(record_id):
        await query.edit_message_text("✅ Data absen berhasil dihapus!", reply_markup=absensi_menu_keyboard())
    else:
        await query.edit_message_text("❌ Data tidak ditemukan.", reply_markup=wrap_keyboard_abs([]))


# ─── DOWNLOAD EXCEL ──────────────────────────────────────

async def abs_excel_month(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    await query.edit_message_text(
        "📥 Kirim bulan dan tahun\nContoh: `7 2026`",
        parse_mode="Markdown",
        reply_markup=wrap_keyboard_abs([]),
    )
    return ABSENSI_EXCEL


async def abs_excel_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    parts = text.split()
    try:
        bulan = int(parts[0])
        tahun = int(parts[1]) if len(parts) > 1 else datetime.now().year
        if bulan < 1 or bulan > 12:
            raise ValueError
    except (ValueError, IndexError):
        await update.message.reply_text("⚠️ Format: bulan tahun, contoh: `7 2026`", parse_mode="Markdown")
        return ABSENSI_EXCEL

    await update.message.reply_text("⏳ Membuat file Excel...")
    month_str = f"{tahun:04d}-{bulan:02d}"
    members = db.get_members()
    if not members:
        await update.message.reply_text("📭 Belum ada anggota.", reply_markup=absensi_menu_keyboard())
        return ConversationHandler.END

    hari_aktif = db.get_attendance_dates(tahun, bulan)
    if not hari_aktif:
        await update.message.reply_text("📭 Belum ada data absen di bulan ini.", reply_markup=absensi_menu_keyboard())
        return ConversationHandler.END

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = f"Absensi {month_str}"

        headers = ["No", "Nama"]
        for h in hari_aktif:
            headers.append(h[-2:])
        headers += ["Hadir", "Izin", "Persentase"]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for i, m in enumerate(members, 1):
            mid = m["id"] if not hasattr(m, "get") else m.get("id")
            mname = m["name"] if not hasattr(m, "get") else m.get("name")
            row_data = [i, mname]
            hadir = 0
            izin = 0
            for date in hari_aktif:
                record = db.get_absensi_record(mid if mid else m.get("id"), date)
                if record and record.get("status") == "izin":
                    izin += 1
                    row_data.append("I")
                elif record and record.get("check_in"):
                    hadir += 1
                    row_data.append("✓")
                else:
                    row_data.append("-")
            pct = (hadir / len(hari_aktif) * 100) if hari_aktif else 0
            row_data += [hadir, izin, f"{pct:.1f}%"]

            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=i + 1, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 3

        # Save to temp file
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)
        tmp.close()

        # Send file via telegram
        with open(tmp.name, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=f"absensi_{bulan:02d}_{tahun}.xlsx",
                caption=f"📥 Data absensi {bulan:02d}-{tahun}",
            )

        os.unlink(tmp.name)

    except Exception as e:
        await update.message.reply_text(f"⚠️ Gagal membuat Excel: {str(e)}")

    await update.message.reply_text("Pilih menu:", reply_markup=absensi_menu_keyboard())
    return ConversationHandler.END
